"""
Export Excel pour l'analyse cross-référentielle
Génère un fichier Excel multi-feuilles avec graphiques
"""

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from datetime import datetime
import io
from typing import Dict, List

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from src.database import get_db

router = APIRouter()


def apply_header_style(cell):
    """Applique le style d'en-tête"""
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    cell.fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


def apply_data_style(cell, bg_color=None):
    """Applique le style de données"""
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")


def create_overview_sheet(wb: Workbook, db: Session):
    """Feuille 1: Vue d'ensemble"""
    ws = wb.active
    ws.title = "Vue d'Ensemble"

    # En-tête principal
    ws['A1'] = "ANALYSE CROSS-RÉFÉRENTIELLE - VUE D'ENSEMBLE"
    ws.merge_cells('A1:E1')
    ws['A1'].font = Font(bold=True, size=16, color="2E5090")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

    # Date de génération
    ws['A2'] = f"Généré le : {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws.merge_cells('A2:E2')
    ws['A2'].font = Font(italic=True, size=10)
    ws['A2'].alignment = Alignment(horizontal="center")

    # Récupérer les statistiques
    stats_query = text("""
        SELECT
            COUNT(DISTINCT cp.id) as total_pcs,
            COUNT(DISTINCT rcp.requirement_id) as total_requirements,
            COUNT(DISTINCT f.id) as total_frameworks,
            COUNT(*) as total_mappings
        FROM control_point cp
        JOIN requirement_control_point rcp ON rcp.control_point_id = cp.id
        JOIN requirement r ON r.id = rcp.requirement_id
        JOIN framework f ON f.id = r.framework_id
    """)
    stats = db.execute(stats_query).fetchone()

    # PCs cross-référentiels
    cross_ref_query = text("""
        SELECT COUNT(*) FROM (
            SELECT cp.id
            FROM control_point cp
            JOIN requirement_control_point rcp ON rcp.control_point_id = cp.id
            JOIN requirement r ON r.id = rcp.requirement_id
            JOIN framework f ON f.id = r.framework_id
            GROUP BY cp.id
            HAVING COUNT(DISTINCT f.id) > 1
        ) as cross_ref_pcs
    """)
    cross_ref_count = db.execute(cross_ref_query).fetchone()[0]

    # Tableau des KPIs
    row = 4
    ws[f'A{row}'] = "INDICATEUR"
    ws[f'B{row}'] = "VALEUR"
    ws[f'C{row}'] = "DESCRIPTION"
    for col in ['A', 'B', 'C']:
        apply_header_style(ws[f'{col}{row}'])

    kpis = [
        ("Total Points de Contrôle", stats.total_pcs, "Nombre total de PCs uniques"),
        ("Total Exigences", stats.total_requirements, "Nombre total d'exigences liées"),
        ("Total Frameworks", stats.total_frameworks, "Nombre de référentiels avec PCs"),
        ("PCs Cross-Référentiels", cross_ref_count, "PCs partagés entre plusieurs frameworks"),
        ("Taux Cross-Référentiels", f"{(cross_ref_count/stats.total_pcs*100):.1f}%", "% de PCs partagés"),
        ("Taux Déduplication", f"{((stats.total_requirements-stats.total_pcs)/stats.total_requirements*100):.1f}%", "Économie de PCs réalisée"),
        ("PCs Évités", stats.total_requirements - stats.total_pcs, "Nombre de PCs économisés"),
    ]

    row = 5
    for kpi_name, kpi_value, kpi_desc in kpis:
        ws[f'A{row}'] = kpi_name
        ws[f'B{row}'] = kpi_value
        ws[f'C{row}'] = kpi_desc
        for col in ['A', 'B', 'C']:
            apply_data_style(ws[f'{col}{row}'])
        ws[f'B{row}'].alignment = Alignment(horizontal="center", vertical="center")
        ws[f'B{row}'].font = Font(bold=True, size=12)
        row += 1

    # Ajuster les largeurs
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50

    # Ajouter un graphique en camembert
    pie = PieChart()
    pie.title = "Répartition des PCs"
    pie.style = 10

    # Données pour le graphique
    chart_row = row + 2
    ws[f'A{chart_row}'] = "Catégorie"
    ws[f'B{chart_row}'] = "Nombre"
    apply_header_style(ws[f'A{chart_row}'])
    apply_header_style(ws[f'B{chart_row}'])

    ws[f'A{chart_row+1}'] = "PCs Cross-Référentiels"
    ws[f'B{chart_row+1}'] = cross_ref_count
    ws[f'A{chart_row+2}'] = "PCs Mono-Référentiel"
    ws[f'B{chart_row+2}'] = stats.total_pcs - cross_ref_count

    labels = Reference(ws, min_col=1, min_row=chart_row+1, max_row=chart_row+2)
    data = Reference(ws, min_col=2, min_row=chart_row, max_row=chart_row+2)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    ws.add_chart(pie, f"E{chart_row}")


def create_matrix_sheet(wb: Workbook, db: Session):
    """Feuille 2: Matrice de couverture"""
    ws = wb.create_sheet("Matrice de Couverture")

    # En-tête
    ws['A1'] = "MATRICE DE COUVERTURE CROSS-RÉFÉRENTIELLE"
    ws.merge_cells('A1:F1')
    ws['A1'].font = Font(bold=True, size=14, color="2E5090")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

    # Récupérer les frameworks
    frameworks_query = text("""
        SELECT DISTINCT f.id, f.code, f.name
        FROM framework f
        JOIN requirement r ON r.framework_id = f.id
        JOIN requirement_control_point rcp ON rcp.requirement_id = r.id
        WHERE f.is_active = true
        ORDER BY f.code
    """)
    frameworks = db.execute(frameworks_query).fetchall()

    # Créer la matrice
    row = 3
    ws[f'A{row}'] = "Framework"
    col_idx = 2
    for fw in frameworks:
        cell = ws.cell(row=row, column=col_idx)
        cell.value = fw.code
        apply_header_style(cell)
        col_idx += 1
    apply_header_style(ws[f'A{row}'])

    # Remplir la matrice
    for i, source_fw in enumerate(frameworks):
        row = 4 + i
        ws[f'A{row}'] = source_fw.code
        apply_header_style(ws[f'A{row}'])

        for j, target_fw in enumerate(frameworks):
            col_idx = 2 + j
            cell = ws.cell(row=row, column=col_idx)

            if source_fw.code == target_fw.code:
                cell.value = "—"
                apply_data_style(cell, "D3D3D3")
            else:
                # Calculer les PCs partagés
                shared_query = text("""
                    SELECT COUNT(DISTINCT cp.id) as shared_count
                    FROM control_point cp
                    WHERE EXISTS (
                        SELECT 1 FROM requirement_control_point rcp1
                        JOIN requirement r1 ON r1.id = rcp1.requirement_id
                        WHERE rcp1.control_point_id = cp.id
                        AND r1.framework_id = :source_id
                    )
                    AND EXISTS (
                        SELECT 1 FROM requirement_control_point rcp2
                        JOIN requirement r2 ON r2.id = rcp2.requirement_id
                        WHERE rcp2.control_point_id = cp.id
                        AND r2.framework_id = :target_id
                    )
                """)
                result = db.execute(shared_query, {
                    "source_id": str(source_fw.id),
                    "target_id": str(target_fw.id)
                }).fetchone()

                shared_count = result.shared_count if result else 0

                # Total PCs du framework cible
                total_query = text("""
                    SELECT COUNT(DISTINCT cp.id) as total
                    FROM control_point cp
                    JOIN requirement_control_point rcp ON rcp.control_point_id = cp.id
                    JOIN requirement r ON r.id = rcp.requirement_id
                    WHERE r.framework_id = :target_id
                """)
                total_result = db.execute(total_query, {"target_id": str(target_fw.id)}).fetchone()
                total = total_result.total if total_result else 1

                percentage = (shared_count / total * 100) if total > 0 else 0
                cell.value = f"{percentage:.1f}%"

                # Couleur selon le pourcentage
                if percentage >= 15:
                    bg_color = "C6EFCE"  # Vert
                elif percentage >= 10:
                    bg_color = "9BC2E6"  # Bleu
                elif percentage >= 5:
                    bg_color = "FFD966"  # Orange
                else:
                    bg_color = "FFC7CE"  # Rouge

                apply_data_style(cell, bg_color)
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ajuster les largeurs
    ws.column_dimensions['A'].width = 20
    for i in range(len(frameworks)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 15


def create_shared_pcs_sheet(wb: Workbook, db: Session):
    """Feuille 3: Liste des PCs partagés"""
    ws = wb.create_sheet("PCs Partagés")

    # En-tête
    ws['A1'] = "POINTS DE CONTRÔLE CROSS-RÉFÉRENTIELS"
    ws.merge_cells('A1:G1')
    ws['A1'].font = Font(bold=True, size=14, color="2E5090")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

    # En-têtes de colonnes
    headers = ["Code PC", "Nom", "Description", "Criticité", "Nb Frameworks", "Nb Exigences", "Frameworks"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_header_style(cell)

    # Récupérer les PCs partagés
    query = text("""
        SELECT
            cp.code,
            cp.name,
            cp.description,
            cp.criticality_level,
            COUNT(DISTINCT f.id) as nb_frameworks,
            COUNT(DISTINCT rcp.requirement_id) as nb_requirements,
            STRING_AGG(DISTINCT f.code, ', ' ORDER BY f.code) as frameworks
        FROM control_point cp
        JOIN requirement_control_point rcp ON rcp.control_point_id = cp.id
        JOIN requirement r ON r.id = rcp.requirement_id
        JOIN framework f ON f.id = r.framework_id
        GROUP BY cp.id, cp.code, cp.name, cp.description, cp.criticality_level
        HAVING COUNT(DISTINCT f.id) > 1
        ORDER BY COUNT(DISTINCT rcp.requirement_id) DESC
    """)

    pcs = db.execute(query).fetchall()

    row = 4
    for pc in pcs:
        ws.cell(row=row, column=1).value = pc.code
        ws.cell(row=row, column=2).value = pc.name
        ws.cell(row=row, column=3).value = pc.description or ""
        ws.cell(row=row, column=4).value = pc.criticality_level or "N/A"
        ws.cell(row=row, column=5).value = pc.nb_frameworks
        ws.cell(row=row, column=6).value = pc.nb_requirements
        ws.cell(row=row, column=7).value = pc.frameworks

        # Style
        for col_idx in range(1, 8):
            cell = ws.cell(row=row, column=col_idx)
            apply_data_style(cell)

        # Couleur selon criticité
        criticality = (pc.criticality_level or "").lower()
        if criticality == "critical":
            ws.cell(row=row, column=4).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif criticality == "high":
            ws.cell(row=row, column=4).fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        elif criticality == "medium":
            ws.cell(row=row, column=4).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif criticality == "low":
            ws.cell(row=row, column=4).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        # Centre pour les nombres
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=6).alignment = Alignment(horizontal="center", vertical="center")

        row += 1

    # Ajuster les largeurs
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 25


def create_top_pcs_sheet(wb: Workbook, db: Session):
    """Feuille 4: Top PCs réutilisés avec graphique"""
    ws = wb.create_sheet("Top PCs Réutilisés")

    # En-tête
    ws['A1'] = "TOP 10 POINTS DE CONTRÔLE LES PLUS RÉUTILISÉS"
    ws.merge_cells('A1:E1')
    ws['A1'].font = Font(bold=True, size=14, color="2E5090")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

    # En-têtes de colonnes
    headers = ["Code PC", "Nom", "Nb Frameworks", "Nb Exigences", "Frameworks"]
    row = 3
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        apply_header_style(cell)

    # Récupérer le top 10
    query = text("""
        SELECT
            cp.code,
            cp.name,
            COUNT(DISTINCT f.id) as nb_frameworks,
            COUNT(DISTINCT rcp.requirement_id) as nb_requirements,
            STRING_AGG(DISTINCT f.code, ', ' ORDER BY f.code) as frameworks
        FROM control_point cp
        JOIN requirement_control_point rcp ON rcp.control_point_id = cp.id
        JOIN requirement r ON r.id = rcp.requirement_id
        JOIN framework f ON f.id = r.framework_id
        GROUP BY cp.id, cp.code, cp.name
        ORDER BY COUNT(DISTINCT rcp.requirement_id) DESC
        LIMIT 10
    """)

    top_pcs = db.execute(query).fetchall()

    row = 4
    for pc in top_pcs:
        ws.cell(row=row, column=1).value = pc.code
        ws.cell(row=row, column=2).value = pc.name
        ws.cell(row=row, column=3).value = pc.nb_frameworks
        ws.cell(row=row, column=4).value = pc.nb_requirements
        ws.cell(row=row, column=5).value = pc.frameworks

        for col_idx in range(1, 6):
            cell = ws.cell(row=row, column=col_idx)
            apply_data_style(cell)

        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center", vertical="center")

        row += 1

    # Ajuster les largeurs
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30

    # Ajouter un graphique en barres
    chart = BarChart()
    chart.type = "col"
    chart.title = "Nombre d'exigences par PC"
    chart.x_axis.title = "Points de Contrôle"
    chart.y_axis.title = "Nombre d'exigences"
    chart.style = 10

    data = Reference(ws, min_col=4, min_row=3, max_row=3+len(top_pcs))
    categories = Reference(ws, min_col=1, min_row=4, max_row=3+len(top_pcs))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    ws.add_chart(chart, "G3")


def create_statistics_sheet(wb: Workbook, db: Session):
    """Feuille 5: Statistiques détaillées"""
    ws = wb.create_sheet("Statistiques")

    # En-tête
    ws['A1'] = "STATISTIQUES DÉTAILLÉES"
    ws.merge_cells('A1:D1')
    ws['A1'].font = Font(bold=True, size=14, color="2E5090")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

    # Distribution par nombre de frameworks
    ws['A3'] = "DISTRIBUTION DES PCs PAR NOMBRE DE FRAMEWORKS"
    ws.merge_cells('A3:D3')
    ws['A3'].font = Font(bold=True, size=12)

    row = 4
    ws[f'A{row}'] = "Nb Frameworks"
    ws[f'B{row}'] = "Nb PCs"
    ws[f'C{row}'] = "Pourcentage"
    for col in ['A', 'B', 'C']:
        apply_header_style(ws[f'{col}{row}'])

    # Requête pour la distribution
    dist_query = text("""
        SELECT
            COUNT(DISTINCT f.id) as nb_frameworks,
            COUNT(*) as nb_pcs
        FROM control_point cp
        JOIN requirement_control_point rcp ON rcp.control_point_id = cp.id
        JOIN requirement r ON r.id = rcp.requirement_id
        JOIN framework f ON f.id = r.framework_id
        GROUP BY cp.id
        ORDER BY nb_frameworks
    """)

    distribution = db.execute(dist_query).fetchall()

    # Calculer le total
    total_pcs = sum(d.nb_pcs for d in distribution)

    # Grouper par nb_frameworks
    from collections import defaultdict
    grouped = defaultdict(int)
    for d in distribution:
        grouped[d.nb_frameworks] += 1

    row = 5
    for nb_fw in sorted(grouped.keys()):
        count = grouped[nb_fw]
        percentage = (count / total_pcs * 100) if total_pcs > 0 else 0

        ws[f'A{row}'] = nb_fw
        ws[f'B{row}'] = count
        ws[f'C{row}'] = f"{percentage:.1f}%"

        for col in ['A', 'B', 'C']:
            apply_data_style(ws[f'{col}{row}'])
            ws[f'{col}{row}'].alignment = Alignment(horizontal="center", vertical="center")

        row += 1

    # Graphique
    chart = PieChart()
    chart.title = "Distribution des PCs"
    chart.style = 10

    data = Reference(ws, min_col=2, min_row=4, max_row=row-1)
    labels = Reference(ws, min_col=1, min_row=5, max_row=row-1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)

    ws.add_chart(chart, "E4")

    # Ajuster largeurs
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15


@router.get("/export")
async def export_cross_referentials(db: Session = Depends(get_db)):
    """
    Exporter l'analyse cross-référentielle en Excel
    """
    if not OPENPYXL_AVAILABLE:
        return {"error": "openpyxl n'est pas installé"}

    try:
        # Créer le workbook
        wb = Workbook()

        # Créer toutes les feuilles
        create_overview_sheet(wb, db)
        create_matrix_sheet(wb, db)
        create_shared_pcs_sheet(wb, db)
        create_top_pcs_sheet(wb, db)
        create_statistics_sheet(wb, db)

        # Sauvegarder dans un buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Générer le nom de fichier
        filename = f"cross_referentiels_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # Retourner le fichier
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    except Exception as e:
        print(f"Error in export: {e}")
        import traceback
        traceback.print_exc()
        return {"error": str(e)}
