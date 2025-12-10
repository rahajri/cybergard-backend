# src/api/v1/frameworks.py

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, status, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text, func
from typing import List, Any
import csv
import io
import pandas as pd
import os
from io import BytesIO

# ✅ REDIS CACHE
from src.utils.redis_manager import cache_result

from datetime import datetime
from typing import List, Dict, Any, Optional
from uuid import UUID
import json
from ...services.domain_import_service import DomainImportService 
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ openpyxl non installé - fonctionnalité Excel désactivée")

from ...database import SessionLocal
from ...database import get_db
from src.services.csv_import import import_csv_to_database
from ...services.domain_import_service import DomainImportService
from ...models.audit import Framework, Domain, Requirement
from ...database import get_db

# --- imports DB (robustes) ---
try:
    from src.database import SessionLocal, get_db
except Exception:  # fallback si exécution différente
    from ...database import SessionLocal, get_db  # type: ignore

import logging

logger = logging.getLogger(__name__)

# Ajoutez ces imports après vos imports existants
try:
    from ...services.cross_referential_service import CrossReferentialMappingService
    from ...services.coverage_service import CrossReferentialCoverageService
    CROSS_REFERENTIAL_AVAILABLE = True
except ImportError:
    CROSS_REFERENTIAL_AVAILABLE = False
    print("Services cross-référentiel non disponibles")

router = APIRouter()

def _run_embeddings_in_new_session(framework_id: str):
    # Pas besoin d’ouvrir une session ici : la fonction gère sa propre session
    from ...services.embedding_service import generate_embeddings_for_framework
    generate_embeddings_for_framework(framework_id)


@router.post("/upload")
async def upload_excel_referentiel(
    file: UploadFile = File(...),
    framework_info: str = Form(...),  # JSON des métadonnées
    db: Session = Depends(get_db),
    background: BackgroundTasks = None,  # ✅ pour tâches async
):
    """Upload Excel avec hiérarchie domain"""
    
    # Validation fichier
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Fichier Excel requis (.xlsx ou .xls)")
    
    try:
        # Parser métadonnées
        import json
        metadata = json.loads(framework_info)
        
        # Lire Excel
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        
        # Nettoyer colonnes
        df.columns = df.columns.str.strip()
        
        # Valider colonnes minimales
        required = ['domaine', 'code_officiel', 'titre', 'description']
        missing = [col for col in required if col not in df.columns]
        if missing:
            raise HTTPException(
                status_code=400,
                detail=f"Colonnes manquantes: {', '.join(missing)}"
            )
        
        # Import via service
        from ...services.domain_import_service import DomainImportService
        
        service = DomainImportService(db)
        stats = service.import_excel(df, metadata)

        framework_id = stats.get('framework_id')
        requirements_created = int(stats.get('requirements_created', 0))
        has_errors = bool(stats.get('errors'))
        success = (not has_errors) and (requirements_created > 0)

        # ⚙️ (Optionnel) activer de suite si succès
        if success and framework_id:
            db.execute(text("UPDATE framework SET is_active = TRUE WHERE id = :id"), {"id": framework_id})
            db.commit()

            # ✅ CONDITION : générer les embeddings UNIQUEMENT si succès
            try:
                from ...services.embedding_service import generate_embeddings_for_framework
                if background is not None:
                    background.add_task(generate_embeddings_for_framework, framework_id)
                    embedding_status = "started"
                else:
                    generate_embeddings_for_framework(framework_id)
                    embedding_status = "completed"
            except Exception as e:
                embedding_status = f"error: {e}"
        else:
            embedding_status = "skipped"

        return {
            "success": success,
            "framework_id": framework_id,
            "domains_created": stats.get('domains_created', 0),
            "requirements_created": requirements_created,
            "warnings": stats.get('warnings', []),
            "errors": stats.get('errors', []),
            "embeddings": embedding_status,  # 👈 visibilité côté front
            "message": "Import terminé. Génération des embeddings lancée." if success else "Import terminé (embeddings non lancés)."
        }
        
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Métadonnées JSON invalides")
    except Exception as e:
        logger.error(f"Erreur import: {e}")
        raise HTTPException(status_code=500, detail=f"Erreur import: {str(e)}")

@router.get("/")
@cache_result(ttl=1800, key_prefix="frameworks_list")  # ✅ Cache 30 minutes
async def list_frameworks(db: Session = Depends(get_db)):
    """Liste des référentiels importés depuis la base de données avec statistiques d'embeddings"""
    
    try:
        from ...models.audit import Framework
        from sqlalchemy import text
        
        # Requête enrichie avec statistiques d'embeddings
        query = text("""
            SELECT 
                f.id,
                f.code,
                f.name,
                f.version,
                f.publisher,
                f.language,
                f.import_date,
                f.created_at,
                COUNT(DISTINCT r.id) AS requirements_count,
                COUNT(DISTINCT d.id) AS sections_count,          -- ex-« sections » = nb de domaines
                COUNT(DISTINCT re.id) AS embeddings_count
            FROM framework f
            LEFT JOIN requirement r ON r.framework_id = f.id
            LEFT JOIN domain d ON d.framework_id = f.id         -- ⬅️ remplace 'section'
            LEFT JOIN requirement_embeddings re ON re.requirement_id = r.id
            WHERE f.is_active = true
            GROUP BY 
                f.id, f.code, f.name, f.version, f.publisher, 
                f.language, f.import_date, f.created_at
            HAVING 
                COUNT(DISTINCT r.id) > 0
                AND COUNT(DISTINCT re.id) = COUNT(DISTINCT r.id)  -- embeddings « prêts » (100%)
            ORDER BY f.created_at DESC
        """)

        
        result = db.execute(query)
        frameworks_data = []
        
        for row in result:
            embedding_coverage = 0
            if row.requirements_count > 0:
                embedding_coverage = round((row.embeddings_count / row.requirements_count) * 100, 1)
            
            frameworks_data.append({
                "id": str(row.id),
                "code": row.code,
                "name": row.name,
                "version": row.version,
                "publisher": row.publisher,
                "language": row.language,
                "import_date": row.import_date.isoformat() if row.import_date else None,
                "requirements_count": row.requirements_count,
                "sections_count": row.sections_count,
                "embeddings_count": row.embeddings_count,
                "embedding_coverage": embedding_coverage,
                "embedding_status": "complete" if embedding_coverage == 100 else "partial" if embedding_coverage > 0 else "none"
            })
        
        return {
            "frameworks": frameworks_data,
            "total": len(frameworks_data),
            "message": f"{len(frameworks_data)} référentiel(s) importé(s).",
            "summary": {
                "total_requirements": sum(fw["requirements_count"] for fw in frameworks_data),
                "total_embeddings": sum(fw["embeddings_count"] for fw in frameworks_data),
                "frameworks_with_embeddings": len([fw for fw in frameworks_data if fw["embeddings_count"] > 0])
            }
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur récupération frameworks: {str(e)}")

@router.get("/{framework_id}")
@cache_result(ttl=1800, key_prefix="framework_detail")  # ✅ Cache 30min
async def get_framework(framework_id: str, db: Session = Depends(get_db)):
    """Détail d'un référentiel avec ses exigences"""
    
    try:
        from ...models.audit import Framework
        from sqlalchemy import text
        
        framework = db.query(Framework).filter_by(id=framework_id).first()
        
        if not framework:
            raise HTTPException(status_code=404, detail="Référentiel non trouvé")
        
        # ✅ CORRECTION : Utiliser domain au lieu de section
        query = text("""
            SELECT
                r.id,
                r.official_code,
                r.title,
                r.requirement_text as description,
                r.tags,
                r.risk_level as niveau_risque,
                r.compliance_obligation as obligation_conformite,
                r.chapter_path,
                d.code as domain_code,
                COALESCE(
                    (SELECT dt.title
                    FROM domain_title dt
                    WHERE dt.domain_id = d.id
                    AND dt.is_primary = true
                    AND dt.language = 'fr'
                    LIMIT 1),
                    d.code
                ) as domain_title,
                CASE WHEN re.id IS NOT NULL THEN true ELSE false END as has_embedding,
                re.created_at as embedding_created_at
            FROM requirement r
            LEFT JOIN domain d ON r.domain_id = d.id
            LEFT JOIN requirement_embeddings re ON r.id = re.requirement_id
            WHERE r.framework_id = :framework_id
            ORDER BY r.official_code
        """)
        
        result = db.execute(query, {"framework_id": framework_id})
        requirements_data = []
        embeddings_count = 0
        
        for row in result:
            if row.has_embedding:
                embeddings_count += 1
                
            requirements_data.append({
                "id": str(row.id),
                "official_code": row.official_code or "",
                "title": row.title or "",
                "description": row.description or "",
                "domain_code": row.domain_code or "",  # ✅ Corrigé
                "subdomain": "",  # ✅ Vidé (n'existe plus)
                "tags": row.tags if isinstance(row.tags, list) else [],
                "niveau_risque": row.niveau_risque or "",
                "obligation_conformite": row.obligation_conformite or "",
                "chapter_path": row.chapter_path or "",
                "domain_title": row.domain_title or "",
                "has_embedding": row.has_embedding,
                "embedding_created_at": row.embedding_created_at.isoformat() if row.embedding_created_at else None
            })
        
        return {
            "id": str(framework.id),
            "code": framework.code,
            "name": framework.name,
            "version": framework.version,
            "publisher": framework.publisher,
            "language": framework.language,
            "import_date": framework.import_date.isoformat() if framework.import_date else None,
            "description": framework.description or "",
            "is_active": getattr(framework, 'is_active', True),
            "requirements": requirements_data,
            "statistics": {
                "total_requirements": len(requirements_data),
                "embeddings_count": embeddings_count,
                "embedding_coverage": round((embeddings_count / len(requirements_data) * 100), 1) if requirements_data else 0
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur récupération framework: {e}")
        raise HTTPException(status_code=500, detail=f"Erreur récupération framework: {str(e)}")

@router.post("/{framework_id}/generate-embeddings")
async def generate_embeddings_manual(framework_id: str, db: Session = Depends(get_db)):
    """Génération manuelle des embeddings pour un référentiel spécifique"""
    
    try:
        from ...models.audit import Framework
        
        framework = db.query(Framework).filter_by(id=framework_id, is_active=True).first()
        if not framework:
            raise HTTPException(status_code=404, detail="Référentiel non trouvé")
        
        logger.info(f"Début génération embeddings pour {framework.code}...")
        
        # ✅ CORRECTION : La fonction ne prend QUE framework_id
        from ...services.embedding_service import generate_embeddings_for_framework
        
        result = generate_embeddings_for_framework(framework_id)
        
        if not result or 'error' in result:
            error_msg = result.get('error', 'Erreur inconnue') if result else 'Aucun résultat'
            raise HTTPException(status_code=500, detail=f"Erreur génération: {error_msg}")
        
        return {
            "status": "success",
            "message": f"Embeddings générés pour {framework.code}",
            "framework": {
                "id": framework_id,
                "code": framework.code,
                "name": framework.name
            },
            "results": {
                "total_requirements": result.get('total_requirements', 0),
                "embeddings_generated": result.get('embeddings_generated', 0),
                "errors": result.get('errors', 0),
                "model_used": result.get('model_used', 'xlm-roberta-base'),
                "status": result.get('status', 'completed')
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur génération embeddings: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Erreur: {str(e)}")

# ----- Alias legacy : accepte aussi /frameworks/{id}/export -----
@router.get("/frameworks/{framework_id}/export")
async def export_framework_legacy_alias(
    framework_id: str,
    format: str = "xlsx",
    db: Session = Depends(get_db),
):
    # redirige proprement vers l'endpoint correct "/{id}/export"
    return await export_framework(framework_id=framework_id, format=format, db=db)


@router.get("/template/excel/download")
async def download_excel_template():
    """
    Télécharger le template Excel avec hiérarchie domain (0-4 niveaux)
    Inclut des exemples et une feuille d'instructions
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Template_Import"
        
        # ===== EN-TÊTES =====
        headers = [
            "Nom référentiel",      # A - Métadonnée (optionnel, renseigné dans l'UI)
            "Formule",              # B - Métadonnée (auto-généré)
            "domaine",              # C - Niveau 0 (OBLIGATOIRE)
            "domaine_rang1",        # D - Niveau 1 (optionnel)
            "domaine_rang2",        # E - Niveau 2 (optionnel)
            "domaine_rang3",        # F - Niveau 3 (optionnel)
            "domaine_rang4",        # G - Niveau 4 (optionnel)
            "code_officiel",        # H - Code exigence (OBLIGATOIRE)
            "titre",                # I - Titre exigence (OBLIGATOIRE)
            "description",          # J - Texte complet (OBLIGATOIRE)
            "tags",                 # K - Domaine audit alternatif
            "niveau_risque",        # L - low/medium/high/critical
            "obligation_conformite",# M - mandatory/recommended/optional
            "profil"                # N - Profil d'application
        ]
        
        # Style header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        
        # ===== EXEMPLES DE DONNÉES =====
        examples = [
            # Exemple 1 : Niveau 0 uniquement (simple)
            [
                "ISO 27002:2022",  # Nom référentiel (optionnel ici)
                "",                # Formule (auto-généré par le système)
                "Mesures organisationnelles",  # domaine (niveau 0)
                "",                # domaine_rang1 (vide)
                "",                # domaine_rang2 (vide)
                "",                # domaine_rang3 (vide)
                "",                # domaine_rang4 (vide)
                "05.01",           # code_officiel
                "Politiques de sécurité de l'information",  # titre
                "Il convient de définir une politique de sécurité de l'information et des politiques spécifiques à une thématique, de les faire approuver par la direction, de les publier, de les communiquer et d'en demander confirmation au personnel et aux parties intéressées concernés, ainsi que de les réviser à intervalles planifiés et si des changements significatifs ont lieu.",  # description
                "A.5 Politique de sécurité de l'information",  # tags
                "medium",          # niveau_risque
                "mandatory",       # obligation_conformite
                ""                 # profil
            ],
            
            # Exemple 2 : 2 niveaux de hiérarchie
            [
                "",
                "",
                "Mesures technologiques",     # Niveau 0
                "Contrôle d'accès",           # Niveau 1
                "",
                "",
                "",
                "08.01",
                "Gestion des accès utilisateurs",
                "Les droits d'accès aux informations et autres actifs associés doivent être attribués, vérifiés, modifiés et supprimés conformément aux politiques spécifiques de l'organisation relatives au contrôle d'accès et aux règles métier.",
                "A.8 Gestion des actifs",
                "high",
                "mandatory",
                ""
            ],
            
            # Exemple 3 : 3 niveaux de hiérarchie
            [
                "",
                "",
                "Mesures physiques",           # Niveau 0
                "Sécurité périmétrique",       # Niveau 1
                "Contrôle accès bâtiment",     # Niveau 2
                "",
                "",
                "07.01",
                "Périmètres de sécurité physique",
                "Des périmètres de sécurité doivent être définis et utilisés pour protéger les zones contenant des informations et autres actifs associés sensibles ou critiques.",
                "A.7 Sécurité physique",
                "high",
                "mandatory",
                ""
            ],
            
            # Exemple 4 : Même domaine niveau 0, différent niveau 1
            [
                "",
                "",
                "Mesures technologiques",      # Niveau 0 (réutilisé)
                "Cryptographie",               # Niveau 1 (nouveau)
                "",
                "",
                "",
                "08.24",
                "Utilisation de la cryptographie",
                "Des règles d'utilisation efficace de la cryptographie, y compris pour la gestion des clés cryptographiques, doivent être définies et mises en œuvre.",
                "A.8 Gestion des actifs",
                "high",
                "mandatory",
                ""
            ]
        ]
        
        for row_data in examples:
            ws.append(row_data)
        
        # ===== LARGEURS DE COLONNES =====
        column_widths = [20, 20, 30, 25, 25, 25, 25, 15, 40, 60, 35, 15, 20, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        # Fixer hauteur ligne header
        ws.row_dimensions[1].height = 30
        
        # ===== FEUILLE INSTRUCTIONS =====
        ws_instructions = wb.create_sheet("📖 Instructions")
        
        instructions = [
            ["GUIDE D'UTILISATION DU TEMPLATE", ""],
            ["", ""],
            ["1. COLONNES OBLIGATOIRES", ""],
            ["━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", ""],
            ["domaine", "Domaine racine (niveau 0) - OBLIGATOIRE pour chaque exigence"],
            ["code_officiel", "Code officiel de l'exigence (ex: 05.01, A.5.1)"],
            ["titre", "Titre court de l'exigence"],
            ["description", "Texte complet de l'exigence"],
            ["", ""],
            ["2. HIÉRARCHIE DES DOMAINES (optionnel)", ""],
            ["━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", ""],
            ["domaine_rang1", "Sous-domaine niveau 1 (optionnel)"],
            ["domaine_rang2", "Sous-domaine niveau 2 (optionnel)"],
            ["domaine_rang3", "Sous-domaine niveau 3 (optionnel)"],
            ["domaine_rang4", "Sous-domaine niveau 4 (optionnel)"],
            ["", ""],
            ["💡 ASTUCE", "La hiérarchie se construit automatiquement !"],
            ["", "Si plusieurs lignes ont le même domaine, il sera réutilisé."],
            ["", ""],
            ["3. COLONNES COMPLÉMENTAIRES", ""],
            ["━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", ""],
            ["tags", "Domaine d'audit alternatif (ex: A.5, A.6...) pour classification ISO 27002"],
            ["niveau_risque", "Criticité : low, medium, high, critical (défaut: medium)"],
            ["obligation_conformite", "Type : mandatory, recommended, optional (défaut: mandatory)"],
            ["", ""],
            ["4. MÉTADONNÉES (gérées dans l'interface)", ""],
            ["━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", ""],
            ["Nom référentiel", "Renseigné dans l'interface d'import (ex: ISO 27002:2022)"],
            ["Formule", "Généré automatiquement (ex: 27002_2022_FR_)"],
            ["", ""],
            ["5. RÈGLES IMPORTANTES", ""],
            ["━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", ""],
            ["✅", "Le domaine (niveau 0) est OBLIGATOIRE pour chaque exigence"],
            ["✅", "Les sous-domaines (rang1-4) sont optionnels selon votre structure"],
            ["✅", "Les domaines identiques sont automatiquement réutilisés"],
            ["✅", "Si aucun domaine n'est défini, l'exigence sera classée dans 'Autre'"],
            ["✅", "Les tags permettent une classification alternative (ex: 14 domaines ISO 27002)"],
            ["", ""],
            ["6. EXEMPLE D'UTILISATION", ""],
            ["━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", ""],
            ["Ligne 1", "domaine='Mesures organisationnelles', code='05.01' → Niveau 0"],
            ["Ligne 2", "domaine='Mesures techniques', rang1='Accès', code='08.01' → 2 niveaux"],
            ["Ligne 3", "domaine='Mesures techniques', rang1='Crypto', code='08.24' → Réutilise niveau 0"],
            ["", ""],
            ["📞 SUPPORT", ""],
            ["━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", ""],
            ["En cas de problème", "Consultez les logs d'import pour identifier les erreurs"],
            ["", f"Template généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}"]
        ]
        
        # Style instructions
        title_font = Font(bold=True, size=14, color="1F4E78")
        section_font = Font(bold=True, size=11, color="2F5496")
        
        for row_idx, row_data in enumerate(instructions, 1):
            ws_instructions.append(row_data)
            
            # Style première colonne
            cell_a = ws_instructions.cell(row=row_idx, column=1)
            cell_b = ws_instructions.cell(row=row_idx, column=2)
            
            if row_data[0] == "GUIDE D'UTILISATION DU TEMPLATE":
                cell_a.font = title_font
            elif "━━━" in row_data[0] or row_data[0].startswith(("1.", "2.", "3.", "4.", "5.", "6.")):
                cell_a.font = section_font
            
            # Couleurs spéciales
            if row_data[0] == "✅":
                cell_a.font = Font(color="00B050", size=12)
            elif row_data[0] in ["💡 ASTUCE", "📞 SUPPORT"]:
                cell_a.font = Font(bold=True, size=11, color="C65911")
        
        # Largeurs colonnes instructions
        ws_instructions.column_dimensions['A'].width = 30
        ws_instructions.column_dimensions['B'].width = 80
        
        # ===== RETOUR FICHIER =====
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger.info("✅ Template Excel généré avec succès")
        
        return StreamingResponse(
            excel_buffer,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                "Content-Disposition": "attachment; filename=template_import_referentiel.xlsx"
            }
        )
        
    except Exception as e:
        logger.error(f"❌ Erreur génération template : {e}")
        raise HTTPException(status_code=500, detail=f"Erreur génération template : {str(e)}")

# Remplacer TOUTES les définitions existantes de /{framework_id}/export et /frameworks/{framework_id}/export par celle-ci

@router.get("/{framework_id}/export")
async def export_framework(framework_id: str, format: str = "xlsx", db: Session = Depends(get_db)):
    """
    Export d'un référentiel avec structure identique au template d'import.
    Détection automatique de la profondeur hiérarchique.
    """
    try:
        # 1️⃣ Métadonnées du framework
        fw_query = text("""
            SELECT code, name, version, publisher, language
            FROM framework
            WHERE id = :fw_id
        """)
        fw_row = db.execute(fw_query, {"fw_id": framework_id}).mappings().first()

        if not fw_row:
            raise HTTPException(status_code=404, detail="Référentiel non trouvé")

        # 2️⃣ Détection profondeur hiérarchique
        max_level_query = text("""
            SELECT COALESCE(MAX(level), 0) as max_level
            FROM domain
            WHERE framework_id = :fw_id
        """)
        max_level_row = db.execute(max_level_query, {"fw_id": framework_id}).mappings().first()
        max_level = max_level_row['max_level'] if max_level_row else 0
        
        logger.info(f"📊 Export {fw_row['code']} - Hiérarchie: {max_level + 1} niveaux")

        # 3️⃣ Récupérer exigences avec leur chemin hiérarchique complet
        requirements_query = text("""
            SELECT 
                r.id,
                r.official_code,
                r.title,
                r.requirement_text,
                r.tags,
                r.risk_level,
                r.compliance_obligation,
                r.domain_id
            FROM requirement r
            WHERE r.framework_id = :fw_id
            ORDER BY r.official_code NULLS LAST
        """)
        
        requirements = db.execute(requirements_query, {"fw_id": framework_id}).mappings().all()

        # 4️⃣ Pour chaque exigence, récupérer son chemin hiérarchique
        export_data = []
        
        for req in requirements:
            row_data = {}
            
            # Si l'exigence a un domain_id, reconstruire la hiérarchie
            if req['domain_id']:
                # Récupérer le chemin complet depuis le domain vers la racine
                path_query = text("""
                    WITH RECURSIVE domain_hierarchy AS (
                        -- Nœud actuel
                        SELECT 
                            d.id,
                            d.parent_id,
                            d.level,
                            COALESCE(
                                (SELECT dt.title FROM domain_title dt 
                                 WHERE dt.domain_id = d.id AND dt.is_primary = true AND dt.language = 'fr' 
                                 LIMIT 1),
                                d.code
                            ) as title
                        FROM domain d
                        WHERE d.id = :domain_id
                        
                        UNION ALL
                        
                        -- Parents récursifs
                        SELECT 
                            d.id,
                            d.parent_id,
                            d.level,
                            COALESCE(
                                (SELECT dt.title FROM domain_title dt 
                                 WHERE dt.domain_id = d.id AND dt.is_primary = true AND dt.language = 'fr' 
                                 LIMIT 1),
                                d.code
                            ) as title
                        FROM domain d
                        INNER JOIN domain_hierarchy dh ON d.id = dh.parent_id
                    )
                    SELECT level, title
                    FROM domain_hierarchy
                    ORDER BY level
                """)
                
                path_result = db.execute(path_query, {"domain_id": req['domain_id']}).mappings().all()
                
                # Remplir les colonnes domaine selon le niveau
                for domain_row in path_result:
                    level = domain_row['level']
                    title = domain_row['title']
                    
                    if level == 0:
                        row_data['domaine'] = title
                    else:
                        row_data[f'domaine_rang{level}'] = title
            
            # Remplir les colonnes vides pour les niveaux non utilisés
            if 'domaine' not in row_data:
                row_data['domaine'] = ""
            
            for i in range(1, max_level + 1):
                col_name = f'domaine_rang{i}'
                if col_name not in row_data:
                    row_data[col_name] = ""
            
            # Informations de l'exigence
            row_data['code_officiel'] = req['official_code'] or ""
            row_data['titre'] = req['title'] or ""
            row_data['description'] = req['requirement_text'] or ""
            
            # Tags
            tags_value = req['tags']
            if isinstance(tags_value, list):
                row_data['tags'] = ",".join([str(t) for t in tags_value])
            elif isinstance(tags_value, str):
                row_data['tags'] = tags_value
            else:
                row_data['tags'] = ""
            
            row_data['niveau_risque'] = req['risk_level'] or ""
            row_data['obligation_conformite'] = req['compliance_obligation'] or ""
            row_data['profil'] = ""
            
            export_data.append(row_data)

        # 5️⃣ Définir les headers
        headers = ['domaine']
        for i in range(1, max_level + 1):
            headers.append(f'domaine_rang{i}')
        headers.extend(['code_officiel', 'titre', 'description', 'tags', 'niveau_risque', 'obligation_conformite', 'profil'])

        base_filename = fw_row['code']

        # 6️⃣ Génération Excel
        if format.lower() == "xlsx":
            if not OPENPYXL_AVAILABLE:
                raise HTTPException(status_code=500, detail="Module openpyxl non disponible")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Export"

            # Headers stylisés
            ws.append(headers)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Données
            for item in export_data:
                ws.append([item.get(h, "") for h in headers])

            # Largeurs colonnes
            for col_idx, header in enumerate(headers, 1):
                if header.startswith('domaine'):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 30
                elif header == 'code_officiel':
                    ws.column_dimensions[get_column_letter(col_idx)].width = 15
                elif header == 'titre':
                    ws.column_dimensions[get_column_letter(col_idx)].width = 40
                elif header == 'description':
                    ws.column_dimensions[get_column_letter(col_idx)].width = 60
                elif header == 'tags':
                    ws.column_dimensions[get_column_letter(col_idx)].width = 35
                else:
                    ws.column_dimensions[get_column_letter(col_idx)].width = 18

            ws.freeze_panes = 'A2'

            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            
            return StreamingResponse(
                bio,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{base_filename}.xlsx"'}
            )

        # CSV
        elif format.lower() == "csv":
            output = io.StringIO()
            if export_data:
                writer = csv.DictWriter(output, fieldnames=headers)
                writer.writeheader()
                writer.writerows(export_data)
            
            return StreamingResponse(
                io.BytesIO(output.getvalue().encode("utf-8")),
                media_type="text/csv",
                headers={"Content-Disposition": f'attachment; filename="{base_filename}.csv"'}
            )

        # JSON
        elif format.lower() == "json":
            payload = {
                "framework": {
                    "code": fw_row["code"],
                    "name": fw_row["name"],
                    "version": fw_row["version"],
                    "hierarchy_depth": max_level + 1
                },
                "requirements": export_data
            }
            
            return StreamingResponse(
                io.BytesIO(json.dumps(payload, indent=2, ensure_ascii=False).encode("utf-8")),
                media_type="application/json",
                headers={"Content-Disposition": f'attachment; filename="{base_filename}.json"'}
            )

        else:
            raise HTTPException(status_code=400, detail="Format non supporté")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erreur export: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Erreur export: {str(e)}")
    
@router.get("/{framework_id}/embeddings/stats")
async def get_embeddings_stats(framework_id: str, db: Session = Depends(get_db)):
    """Statistiques détaillées des embeddings pour un référentiel"""
    
    try:
        from ...models.audit import Framework
        from sqlalchemy import text
        
        framework = db.query(Framework).filter_by(id=framework_id, is_active=True).first()
        if not framework:
            raise HTTPException(status_code=404, detail="Référentiel non trouvé")
        
        # Statistiques détaillées des embeddings
        stats_query = text("""
            SELECT 
                COUNT(r.id) as total_requirements,
                COUNT(re.id) as total_embeddings,
                COUNT(CASE WHEN r.domain IS NOT NULL THEN 1 END) as requirements_with_domain,
                COUNT(CASE WHEN r.tags IS NOT NULL AND r.tags != '[]' THEN 1 END) as requirements_with_tags,
                r.domain,
                COUNT(*) as domain_count,
                AVG(CASE WHEN re.id IS NOT NULL THEN 1.0 ELSE 0.0 END) as embedding_coverage
            FROM requirement r
            LEFT JOIN requirement_embeddings re ON r.id = re.requirement_id
            WHERE r.framework_id = :framework_id
            GROUP BY r.domain
            ORDER BY domain_count DESC
        """)
        
        result = db.execute(stats_query, {"framework_id": framework_id})
        domain_stats = []
        total_requirements = 0
        total_embeddings = 0
        
        for row in result:
            domain_stats.append({
                "domain": row.domain or "Non spécifié",
                "requirements_count": row.domain_count,
                "embedding_coverage": round(row.embedding_coverage * 100, 1) if row.embedding_coverage else 0
            })
            total_requirements += row.domain_count
            if row.embedding_coverage:
                total_embeddings += int(row.domain_count * row.embedding_coverage)
        
        # Vérifier s'il y a des embeddings récents
        recent_embeddings_query = text("""
            SELECT COUNT(*) as recent_count
            FROM requirement_embeddings re
            JOIN requirement r ON re.requirement_id = r.id
            WHERE r.framework_id = :framework_id
            AND re.created_at > NOW() - INTERVAL '1 day'
        """)
        
        recent_result = db.execute(recent_embeddings_query, {"framework_id": framework_id})
        recent_embeddings = recent_result.fetchone().recent_count if recent_result else 0
        
        global_coverage = 0
        if total_requirements > 0:
            global_coverage = round((total_embeddings / total_requirements) * 100, 1)
        
        return {
            "framework": {
                "id": framework_id,
                "code": framework.code,
                "name": framework.name
            },
            "global_stats": {
                "total_requirements": total_requirements,
                "total_embeddings": total_embeddings,
                "global_coverage": global_coverage,
                "recent_embeddings": recent_embeddings,
                "status": "complete" if global_coverage == 100 else "partial" if global_coverage > 0 else "none"
            },
            "domain_breakdown": domain_stats
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur récupération stats embeddings: {str(e)}")
    

# AJOUTS À FAIRE dans votre frameworks.py existant
# Copiez-collez ces endpoints à la fin de votre fichier

# 1. NOUVEAU ENDPOINT : Upload avec cross-référentiel
@router.post("/upload-cross")
async def upload_csv_cross_referentiel(
    file: UploadFile = File(...),
    code_referentiel: str = Form(...),
    nom_referentiel: str = Form(...),
    version: str = Form("1.0"),
    editeur: str = Form(""),
    langue: str = Form("fr"),
    enable_cross_mapping: bool = Form(True),
    similarity_threshold: float = Form(0.75),
    db: Session = Depends(get_db)
):
    """Upload CSV avec détection automatique cross-référentiel"""
    
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Le fichier doit être au format CSV")
    
    try:
        # Votre logique d'import existante
        contents = await file.read()
        df = pd.read_csv(io.BytesIO(contents))
        df.columns = df.columns.str.strip()

        # Validation des colonnes (votre code existant)
        required_columns = ['code_officiel', 'titre', 'description', 'chapitre']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise HTTPException(status_code=400, detail=f"Colonnes manquantes: {', '.join(missing_columns)}")
        
        # Import en base (votre fonction existante)
        stats_db = import_csv_to_database(
            df=df,
            code_referentiel=code_referentiel,
            nom_referentiel=nom_referentiel,
            version=version,
            editeur=editeur,
            langue=langue
        )
        
        # NOUVELLE PARTIE : Génération embeddings + cross-mapping
        framework_id = stats_db['framework_id']
        
        # 1. Génération embeddings (comme votre code existant)
        embedding_stats = {"status": "skipped", "embeddings_generated": 0}
        try:
            from ...services.embedding_service import generate_embeddings_for_framework
            embedding_result = generate_embeddings_for_framework(framework_id)
            
            if 'error' not in embedding_result:
                embedding_stats = {
                    "status": "success",
                    "embeddings_generated": embedding_result.get('embeddings_generated', 0),
                    "total_requirements": embedding_result.get('total_requirements', 0)
                }
        except Exception as e:
            embedding_stats = {"status": "error", "message": str(e)}
        
        # 2. NOUVEAU : Détection cross-référentiel
        cross_mapping_stats = {"status": "skipped", "mappings_detected": 0}
        if enable_cross_mapping and embedding_stats.get("embeddings_generated", 0) > 0:
            try:
                from ...services.cross_referential_service import CrossReferentialMappingService
                
                cross_service = CrossReferentialMappingService(db)
                mapping_result = cross_service.detect_cross_mappings(framework_id, similarity_threshold)
                
                cross_mapping_stats = {
                    "status": "success",
                    "mappings_detected": mapping_result.get("mappings_detected", 0),
                    "auto_validated": mapping_result.get("auto_validated", 0),
                    "pending_validation": mapping_result.get("pending_validation", 0),
                    "similarity_threshold": similarity_threshold
                }
                
            except ImportError:
                cross_mapping_stats = {"status": "disabled", "message": "Service cross-référentiel non disponible"}
            except Exception as e:
                cross_mapping_stats = {"status": "error", "message": str(e)}
        
        return {
            "status": "success",
            "message": f"Fichier {file.filename} traité avec cross-référentiel",
            "referentiel": {
                "id": framework_id,
                "code": code_referentiel,
                "nom": nom_referentiel,
                "version": version
            },
            "statistiques": {
                "sections_created": stats_db["sections_created"],
                "requirements_created": stats_db["requirements_created"],
                "embeddings": embedding_stats,
                "cross_mapping": cross_mapping_stats  # NOUVEAU
            }
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur de traitement: {str(e)}")

# 2. NOUVEAU ENDPOINT : Statut cross-référentiel
@router.get("/{framework_id}/cross-status")
async def get_cross_referential_status(framework_id: str, db: Session = Depends(get_db)):
    """Obtenir le statut cross-référentiel d'un framework"""
    
    try:
        from ...models.audit import Framework
        from sqlalchemy import text
        
        framework = db.query(Framework).filter_by(id=framework_id, is_active=True).first()
        if not framework:
            raise HTTPException(status_code=404, detail="Référentiel non trouvé")
        
        # Statistiques cross-référentiel
        cross_stats_query = text("""
            SELECT 
                COUNT(DISTINCT r.id) as total_requirements,
                COUNT(DISTINCT re.id) as total_embeddings,
                COUNT(DISTINCT rm.id) as total_mappings,
                COUNT(DISTINCT CASE WHEN rm.validation_status = 'approved' THEN rm.id END) as validated_mappings,
                COUNT(DISTINCT CASE WHEN rm.validation_status = 'pending' THEN rm.id END) as pending_mappings
            FROM requirement r
            LEFT JOIN requirement_embeddings re ON r.id = re.requirement_id
            LEFT JOIN requirement_mapping rm ON r.id = rm.source_requirement_id
            WHERE r.framework_id = :framework_id
        """)
        
        result = db.execute(cross_stats_query, {"framework_id": framework_id}).fetchone()
        
        # Frameworks équivalents trouvés
        equivalent_frameworks_query = text("""
            SELECT DISTINCT 
                f2.code,
                f2.name,
                COUNT(rm.id) as mappings_count
            FROM requirement r1
            JOIN requirement_mapping rm ON r1.id = rm.source_requirement_id
            JOIN requirement r2 ON rm.target_requirement_id = r2.id
            JOIN framework f2 ON r2.framework_id = f2.id
            WHERE r1.framework_id = :framework_id
            AND rm.validation_status = 'approved'
            GROUP BY f2.id, f2.code, f2.name
            ORDER BY mappings_count DESC
        """)
        
        equivalent_frameworks = []
        for row in db.execute(equivalent_frameworks_query, {"framework_id": framework_id}):
            equivalent_frameworks.append({
                "code": row.code,
                "name": row.name,
                "mappings_count": row.mappings_count
            })
        
        embedding_coverage = 0
        mapping_coverage = 0
        validation_rate = 0
        
        if result.total_requirements > 0:
            embedding_coverage = round((result.total_embeddings / result.total_requirements) * 100, 1)
            if result.total_mappings > 0:
                mapping_coverage = round((result.total_mappings / result.total_requirements) * 100, 1)
                validation_rate = round((result.validated_mappings / result.total_mappings) * 100, 1)
        
        return {
            "framework": {
                "id": framework_id,
                "code": framework.code,
                "name": framework.name
            },
            "cross_referential_status": {
                "total_requirements": result.total_requirements,
                "embeddings_generated": result.total_embeddings,
                "embedding_coverage": embedding_coverage,
                "mappings_detected": result.total_mappings,
                "mapping_coverage": mapping_coverage,
                "validated_mappings": result.validated_mappings,
                "pending_mappings": result.pending_mappings,
                "validation_rate": validation_rate
            },
            "capabilities": {
                "can_generate_questionnaires": result.total_embeddings > 0,
                "can_calculate_coverage": result.validated_mappings > 0,
                "cross_referential_ready": result.validated_mappings > 0
            },
            "equivalent_frameworks": equivalent_frameworks
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur statut cross-référentiel: {str(e)}")

# 3. NOUVEAU ENDPOINT : Mappings en attente
@router.get("/{framework_id}/pending-mappings")
async def get_pending_mappings(
    framework_id: str, 
    limit: int = 20,
    db: Session = Depends(get_db)
):
    """Récupérer les mappings en attente de validation"""
    
    try:
        from ...services.cross_referential_service import CrossReferentialMappingService
        
        cross_service = CrossReferentialMappingService(db)
        mappings = cross_service.get_pending_mappings(framework_id, limit)
        
        return {
            "framework_id": framework_id,
            "pending_mappings": mappings,
            "total_pending": len(mappings),
            "message": f"{len(mappings)} mapping(s) en attente de validation"
        }
        
    except ImportError:
        raise HTTPException(status_code=500, detail="Service cross-référentiel non disponible")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur récupération mappings: {str(e)}")

# 4. NOUVEAU ENDPOINT : Valider un mapping

# backend/src/api/v1/frameworks.py


@router.post("/upload-excel")
async def upload_excel_alias(
    background: BackgroundTasks,  # ✅ paramètre sans défaut AVANT les autres
    file: UploadFile = File(...),
    framework_info: str = Form(...),
    db: Session = Depends(get_db),
):
    try:
        meta = json.loads(framework_info or "{}")

        # Lecture Excel (bytes -> BytesIO) + engine explicite
        content = await file.read()
        df = pd.read_excel(BytesIO(content), engine="openpyxl")
        df.columns = [str(c).strip() for c in df.columns]

        # Importer (framework/domain/domain_title/requirement)
        importer = DomainImportService(db)
        result = importer.import_excel(df, meta)

        # Commit des insertions
        db.commit()

        # Embeddings REQUIREMENTS (fonction de module, en tâche de fond)
        try:
            from ...services.embedding_service import generate_embeddings_for_framework
        except Exception:
            from src.services.embedding_service import generate_embeddings_for_framework  # fallback absolu

        fw_id = result.get("framework_id")
        if fw_id:
            background.add_task(generate_embeddings_for_framework, fw_id)

        # Réponse pour la popup
        return {
            "success": len(result.get("errors", [])) == 0,
            "framework_id": result.get("framework_id"),
            "framework_name": result.get("framework_name"),
            "domains_created": int(result.get("domains_created", 0)),
            "requirements_created": int(result.get("requirements_created", 0)),
            "warnings": result.get("warnings", []),
            "errors": result.get("errors", []),
        }

    except Exception as e:
        db.rollback()
        logger.exception("Upload Excel failed: %s", e)
        raise HTTPException(status_code=500, detail=f"Erreur import: {e}")



@router.post("/mappings/{mapping_id}/validate")
async def validate_cross_mapping(
    mapping_id: str,
    approved: bool,
    rationale: str = None,
    db: Session = Depends(get_db)
):
    """Valider ou rejeter un mapping cross-référentiel"""
    
    try:
        from ...services.cross_referential_service import CrossReferentialMappingService
        
        # Pour l'exemple, utilisateur fictif - dans votre vraie app, récupérez l'utilisateur connecté
        user_id = "00000000-0000-0000-0000-000000000001"
        
        cross_service = CrossReferentialMappingService(db)
        success = cross_service.validate_mapping(mapping_id, user_id, approved, rationale)
        
        if not success:
            raise HTTPException(status_code=500, detail="Échec de la validation")
        
        return {
            "mapping_id": mapping_id,
            "validation_status": "approved" if approved else "rejected",
            "validated_by": user_id,
            "rationale": rationale,
            "message": f"Mapping {'approuvé' if approved else 'rejeté'} avec succès"
        }
        
    except ImportError:
        raise HTTPException(status_code=500, detail="Service cross-référentiel non disponible")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur validation mapping: {str(e)}")

# 5. NOUVEAU ENDPOINT : Test de similarité
@router.post("/test-similarity")
async def test_similarity_search(
    query_text: str,
    framework_id: str = None,
    limit: int = 5,
    db: Session = Depends(get_db)
):
    """Tester la recherche de similarité cross-référentiel"""
    
    try:
        from ...services.embedding_service import RequirementEmbeddingService
        
        embedding_service = RequirementEmbeddingService(db)
        similar_requirements = embedding_service.find_similar_requirements(
            query_text, framework_id, limit=limit
        )
        
        return {
            "query": query_text,
            "framework_filter": framework_id,
            "similar_requirements": similar_requirements,
            "total_found": len(similar_requirements),
            "message": f"Trouvé {len(similar_requirements)} exigence(s) similaire(s)"
        }
        
    except ImportError:
        raise HTTPException(status_code=500, detail="Service embedding non disponible")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur test similarité: {str(e)}")

# 6. NOUVEAU ENDPOINT : Reprocesser les mappings
@router.post("/{framework_id}/reprocess-mappings")
async def reprocess_cross_mappings(
    framework_id: str,
    similarity_threshold: float = 0.75,
    db: Session = Depends(get_db)
):
    """Recalculer les mappings cross-référentiels"""
    
    try:
        from ...models.audit import Framework
        from ...services.cross_referential_service import CrossReferentialMappingService
        from sqlalchemy import text
        
        framework = db.query(Framework).filter_by(id=framework_id, is_active=True).first()
        if not framework:
            raise HTTPException(status_code=404, detail="Référentiel non trouvé")
        
        # Supprimer anciens mappings
        db.execute(text("""
            DELETE FROM requirement_mapping 
            WHERE source_requirement_id IN (
                SELECT id FROM requirement WHERE framework_id = :framework_id
            )
        """), {"framework_id": framework_id})
        
        # Regénérer
        cross_service = CrossReferentialMappingService(db)
        result = cross_service.detect_cross_mappings(framework_id, similarity_threshold)
        
        db.commit()
        
        return {
            "framework": {
                "id": framework_id,
                "code": framework.code,
                "name": framework.name
            },
            "reprocessing_results": {
                "status": "completed",
                "similarity_threshold": similarity_threshold,
                "mappings_detected": result.get("mappings_detected", 0),
                "auto_validated": result.get("auto_validated", 0),
                "pending_validation": result.get("pending_validation", 0)
            },
            "message": f"Mappings recalculés pour {framework.code}"
        }
        
    except ImportError:
        raise HTTPException(status_code=500, detail="Service cross-référentiel non disponible")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur reprocessing: {str(e)}")

# 🔧 CORRECTION COMPLÈTE de l'endpoint admin cross-referential-summary

@router.get("/admin/cross-referential-summary")
async def get_cross_referential_summary(db: Session = Depends(get_db)) -> Dict[str, Any]:
    """
    Admin summary of cross-referential data: global stats and per-framework details.
    Safe version: no hard dependency on service methods; checks table presence defensively.
    """
    try:
        # 1) Frameworks + embeddings/domains/requirements (sans les mappings pour l’instant)
        frameworks_query = text("""
            SELECT 
                f.id,
                f.code,
                f.name,
                f.version,
                f.publisher,
                f.language,
                f.import_date,
                f.is_active,
                COUNT(DISTINCT r.id)  AS requirements_count,
                COUNT(DISTINCT d.id)  AS domains_count,
                COUNT(DISTINCT re.id) AS embeddings_count
            FROM framework f
            LEFT JOIN requirement r          ON f.id = r.framework_id
            LEFT JOIN domain d               ON f.id = d.framework_id
            LEFT JOIN requirement_embeddings re ON r.id = re.requirement_id
            GROUP BY f.id
            ORDER BY f.import_date DESC NULLS LAST
        """)
        frameworks_rows = db.execute(frameworks_query).fetchall()

        # 2) Existe-t-il une table requirement_mapping ?
        mapping_table_exists = db.execute(
            text("SELECT to_regclass('public.requirement_mapping') IS NOT NULL AS exists")
        ).scalar() or False

        # 3) Si la table existe, calculer le nombre de mappings par framework (sans méthode service)
        per_framework_mappings: Dict[str, int] = {}
        total_mappings = 0

        if mapping_table_exists:
            # On compte un mapping si le framework apparaît soit côté source soit côté target
            mappings_query = text("""
                SELECT f.id AS framework_id, COUNT(DISTINCT rm.id) AS mappings_count
                FROM framework f
                JOIN requirement r ON r.framework_id = f.id
                JOIN requirement_mapping rm 
                  ON rm.source_requirement_id = r.id
                  OR rm.target_requirement_id = r.id
                GROUP BY f.id
            """)
            for row in db.execute(mappings_query).fetchall():
                fw_id = str(row.framework_id)
                count = int(row.mappings_count or 0)
                per_framework_mappings[fw_id] = count
                total_mappings += count

        # 4) Construire la liste frameworks + stats globales
        frameworks_list: List[Dict[str, Any]] = []
        total_requirements = 0
        total_embeddings = 0
        total_domains = 0
        frameworks_with_embeddings = 0
        frameworks_with_mappings = 0

        for row in frameworks_rows:
            fw_id = str(row.id)
            requirements_count = int(row.requirements_count or 0)
            embeddings_count   = int(row.embeddings_count or 0)
            domains_count      = int(row.domains_count or 0)

            embedding_coverage = round((embeddings_count / requirements_count * 100), 1) if requirements_count > 0 else 0.0
            mappings_count     = int(per_framework_mappings.get(fw_id, 0)) if mapping_table_exists else 0

            total_requirements += requirements_count
            total_embeddings   += embeddings_count
            total_domains      += domains_count
            if embeddings_count > 0:
                frameworks_with_embeddings += 1
            if mappings_count > 0:
                frameworks_with_mappings += 1

            frameworks_list.append({
                "id": fw_id,
                "code": row.code,
                "name": row.name,
                "version": row.version,
                "publisher": row.publisher,
                "language": row.language,
                "import_date": row.import_date.isoformat() if row.import_date else None,
                "is_active": bool(row.is_active),
                "requirements_count": requirements_count,
                "domains_count": domains_count,
                "embeddings_count": embeddings_count,
                "embedding_coverage": embedding_coverage,
                "mappings_count": mappings_count,
            })

        average_embedding_coverage = round((total_embeddings / total_requirements * 100), 1) if total_requirements > 0 else 0.0

        global_stats = {
            "total_frameworks": len(frameworks_rows),
            "total_requirements": total_requirements,
            "total_embeddings": total_embeddings,
            "total_mappings": total_mappings if mapping_table_exists else 0,
            "average_embedding_coverage": average_embedding_coverage,
            "frameworks_with_mappings": frameworks_with_mappings,
            "frameworks_with_embeddings": frameworks_with_embeddings,
            "total_domains": total_domains,
        }

        return {
            "global_stats": global_stats,
            "frameworks": frameworks_list,
            "timestamp": datetime.now().isoformat(),
            "status": "ok"
        }

    except Exception as e:
        logger.error(f"Error generating cross-referential summary: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Internal error: {str(e)}")


@router.post("/{framework_id}/generate-control-points")
async def generate_control_points(framework_id: str, db: Session = Depends(get_db)):
    """Générer automatiquement des points de contrôle pour un référentiel"""
    
    try:
        from ...services.control_point_generator import generate_control_points_for_framework
        
        # Vérifier que le framework existe
        framework = db.query(Framework).filter_by(id=framework_id, is_active=True).first()
        if not framework:
            raise HTTPException(status_code=404, detail="Référentiel non trouvé")
        
        # Générer les points de contrôle
        result = await generate_control_points_for_framework(framework_id, db)
        
        if "error" in result:
            raise HTTPException(status_code=500, detail=result["error"])
        
        return {
            "status": "success",
            "message": f"Points de contrôle générés pour {framework.code}",
            "framework": {
                "id": framework_id,
                "code": framework.code,
                "name": framework.name
            },
            "generation_results": result
        }
        
    except ImportError:
        raise HTTPException(status_code=500, detail="Service de génération non disponible")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur génération: {str(e)}")
    
@router.get("/frameworks/{framework_id}")
async def get_framework(
    framework_id: UUID,
    db: Session = Depends(get_db)
):
    """
    Récupère les détails complets d'un référentiel.
    """
    framework = db.query(Framework).filter(Framework.id == framework_id).first()
    
    if not framework:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Référentiel {framework_id} introuvable"
        )
    
    # Compter les exigences
    requirements_count = db.query(func.count(Requirement.id))\
        .filter(Requirement.framework_id == framework_id)\
        .scalar()
    
    # Compter les domaines
    domains_count = db.query(func.count(Domain.id))\
        .filter(Domain.framework_id == framework_id)\
        .scalar()
    
    return {
        "id": str(framework.id),
        "code": framework.code,
        "name": framework.name,
        "version": framework.version,
        "publisher": framework.publisher,
        "language": framework.language,
        "description": framework.description,
        "publication_date": framework.publication_date.isoformat() if framework.publication_date else None,
        "import_date": framework.import_date.isoformat() if framework.import_date else None,
        "source_url": framework.source_url,
        "is_active": framework.is_active,
        "formule": framework.formule,
        "requirements_count": requirements_count,
        "domains_count": domains_count,
        "created_at": framework.created_at.isoformat() if hasattr(framework, 'created_at') else None
    }


# ============ RÉCUPÉRER LA HIÉRARCHIE D'UN RÉFÉRENTIEL ============

@router.get("/{framework_id}/hierarchy")
async def get_framework_hierarchy(framework_id: str, db: Session = Depends(get_db)):
    """Récupère la hiérarchie complète d'un référentiel (domaines + exigences)"""
    
    try:
        from ...models.audit import Framework
        from sqlalchemy import text
        
        # Vérifier que le framework existe
        framework = db.query(Framework).filter_by(id=framework_id).first()
        if not framework:
            raise HTTPException(status_code=404, detail="Référentiel non trouvé")
        
        # Récupérer tous les domaines avec leurs titres
        domains_query = text("""
            WITH RECURSIVE domain_tree AS (
                -- Domaines racines (level 0)
                SELECT 
                    d.id,
                    d.framework_id,
                    d.parent_id,
                    d.code,
                    d.code_officiel,
                    d.level,
                    d.sort_index,
                    COALESCE(
                        (SELECT dt.title 
                         FROM domain_title dt 
                         WHERE dt.domain_id = d.id 
                         AND dt.is_primary = true 
                         AND dt.language = 'fr' 
                         LIMIT 1),
                        d.code
                    ) as title,
                    ARRAY[d.id] as path
                FROM domain d
                WHERE d.framework_id = :framework_id
                AND d.parent_id IS NULL
                
                UNION ALL
                
                -- Sous-domaines récursifs
                SELECT 
                    d.id,
                    d.framework_id,
                    d.parent_id,
                    d.code,
                    d.code_officiel,
                    d.level,
                    d.sort_index,
                    COALESCE(
                        (SELECT dt.title 
                         FROM domain_title dt 
                         WHERE dt.domain_id = d.id 
                         AND dt.is_primary = true 
                         AND dt.language = 'fr' 
                         LIMIT 1),
                        d.code
                    ) as title,
                    dt.path || d.id
                FROM domain d
                INNER JOIN domain_tree dt ON d.parent_id = dt.id
            )
            SELECT * FROM domain_tree
            ORDER BY level, sort_index NULLS LAST, code
        """)
        
        domains_result = db.execute(domains_query, {"framework_id": framework_id}).fetchall()
        
        # Récupérer toutes les exigences du référentiel
        requirements_query = text("""
            SELECT 
                r.id,
                r.domain_id,
                r.official_code,
                r.title,
                r.requirement_text as description,
                r.tags,
                r.risk_level as niveau_risque,
                r.compliance_obligation as obligation_conformite
            FROM requirement r
            WHERE r.framework_id = :framework_id
            ORDER BY r.official_code NULLS LAST
        """)
        
        requirements_result = db.execute(requirements_query, {"framework_id": framework_id}).fetchall()
        
        # Construire un dictionnaire domain_id -> [requirements]
        requirements_by_domain = {}
        for req in requirements_result:
            domain_id = str(req.domain_id) if req.domain_id else None
            if domain_id:
                if domain_id not in requirements_by_domain:
                    requirements_by_domain[domain_id] = []
                requirements_by_domain[domain_id].append({
                    "id": str(req.id),
                    "official_code": req.official_code or "",
                    "title": req.title or "",
                    "description": req.description or "",
                    "tags": req.tags if isinstance(req.tags, str) else "",
                    "niveau_risque": req.niveau_risque or "",
                    "obligation_conformite": req.obligation_conformite or ""
                })
        
        # Construire l'arbre hiérarchique
        domains_dict = {}
        for domain in domains_result:
            domain_id = str(domain.id)
            domains_dict[domain_id] = {
                "id": domain_id,
                "code": domain.code or "",
                "code_officiel": domain.code_officiel or "",
                "title": domain.title or domain.code or "",
                "level": domain.level,
                "parent_id": str(domain.parent_id) if domain.parent_id else None,
                "children": [],
                "requirements": requirements_by_domain.get(domain_id, [])
            }
        
        # Construire la hiérarchie parent-enfant
        root_nodes = []
        for domain_id, domain_data in domains_dict.items():
            parent_id = domain_data["parent_id"]
            if parent_id is None:
                root_nodes.append(domain_data)
            else:
                if parent_id in domains_dict:
                    domains_dict[parent_id]["children"].append(domain_data)
        
        # Trier récursivement
        def sort_tree(nodes):
            for node in nodes:
                node["children"] = sorted(node["children"], key=lambda x: (x["level"], x["code"]))
                if node["children"]:
                    sort_tree(node["children"])
            return sorted(nodes, key=lambda x: (x["level"], x["code"]))
        
        hierarchy = sort_tree(root_nodes)
        
        return {
            "framework_id": framework_id,
            "framework_code": framework.code,
            "framework_name": framework.name,
            "hierarchy": hierarchy,
            "statistics": {
                "total_domains": len(domains_dict),
                "root_domains": len(root_nodes),
                "total_requirements": len(requirements_result)
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Erreur récupération hiérarchie: {e}")
        raise HTTPException(status_code=500, detail=f"Erreur récupération hiérarchie: {str(e)}")

@router.patch("/{framework_id}/toggle-active")
async def toggle_framework_active(
    framework_id: UUID,
    payload: Dict[str, bool],
    db: Session = Depends(get_db)
):
    """
    Active ou désactive un référentiel.
    Seuls les référentiels actifs sont utilisables pour générer des questionnaires.
    """
    # ✅ Import local pour éviter NameError
    from ...models.audit import Framework

    framework = db.query(Framework).filter(Framework.id == framework_id).first()
    if not framework:
        raise HTTPException(status_code=404, detail=f"Référentiel {framework_id} introuvable")

    new_status = payload.get("is_active")
    if new_status is None:
        raise HTTPException(status_code=400, detail="Le champ 'is_active' est requis")

    old_status = framework.is_active
    framework.is_active = bool(new_status)

    try:
        db.commit()
        db.refresh(framework)
        return {
            "success": True,
            "framework_id": str(framework.id),
            "framework_code": framework.code,
            "old_status": old_status,
            "new_status": framework.is_active,
            "message": f"Référentiel {'activé' if framework.is_active else 'désactivé'} avec succès"
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur lors de la mise à jour : {str(e)}")



# ============ SUPPRESSION AVEC CASCADE ============

# src/api/v1/frameworks.py (remplace entièrement la route delete)


@router.delete("/{framework_id}")
async def delete_framework(framework_id: str, db: Session = Depends(get_db)):
    """
    Supprime un référentiel et toutes ses données associées :
    - requirement_mapping (source/target) si la table existe
    - requirement_embeddings si la table existe
    - requirement
    - domain (du plus profond au plus haut)
    - framework
    """
    try:
        from ...models.audit import Framework
        from sqlalchemy import text

        # Existence
        fw = db.query(Framework).filter(Framework.id == framework_id).first()
        if not fw:
            raise HTTPException(status_code=404, detail="Référentiel introuvable")

        # Tables optionnelles
        has_mapping_tbl = db.execute(
            text("SELECT to_regclass('public.requirement_mapping') IS NOT NULL")
        ).scalar() or False

        has_embeddings_tbl = db.execute(
            text("SELECT to_regclass('public.requirement_embeddings') IS NOT NULL")
        ).scalar() or False

        # 1) MAPPINGS liés à ce framework (source OU target)
        if has_mapping_tbl:
            db.execute(text("""
                DELETE FROM requirement_mapping rm
                USING requirement r
                WHERE rm.source_requirement_id = r.id
                  AND r.framework_id = :fid
            """), {"fid": framework_id})

            db.execute(text("""
                DELETE FROM requirement_mapping rm
                USING requirement r
                WHERE rm.target_requirement_id = r.id
                  AND r.framework_id = :fid
            """), {"fid": framework_id})

        # 2) EMBEDDINGS des exigences de ce framework
        if has_embeddings_tbl:
            db.execute(text("""
                DELETE FROM requirement_embeddings e
                USING requirement r
                WHERE e.requirement_id = r.id
                  AND r.framework_id = :fid
            """), {"fid": framework_id})

        # 3) REQUIREMENTS du framework (framework_id est VARCHAR(36))
        db.execute(text("""
            DELETE FROM requirement
            WHERE framework_id = :fid
        """), {"fid": framework_id})

        # 4) DOMAINS du plus profond au plus haut (si pas de cascade sur parent_id)
        max_level = db.execute(text("""
            SELECT COALESCE(MAX(level), 0)
            FROM domain
            WHERE framework_id = :fid
        """), {"fid": framework_id}).scalar() or 0

        for lvl in range(int(max_level), -1, -1):
            db.execute(text("""
                DELETE FROM domain
                WHERE framework_id = :fid
                  AND level = :lvl
            """), {"fid": framework_id, "lvl": lvl})

        # 5) FRAMEWORK (probablement aussi VARCHAR(36))
        db.execute(text("""
            DELETE FROM framework
            WHERE id = :fid
        """), {"fid": framework_id})

        db.commit()
        return {"status": "deleted", "framework_id": framework_id}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erreur suppression: {e}")



# --- dans src/api/v1/frameworks.py ---

@router.get(
    "/{framework_id}/domains",
    response_model=List[Dict[str, Any]],
    summary="Récupérer les domaines d'un référentiel"
)
def get_framework_domains(
    framework_id: str,
    db: Session = Depends(get_db)
):
    """
    Récupère la liste hiérarchique des domaines pour un référentiel donné.
    
    Retourne :
    - Liste des domaines de niveau 0 (racines)
    - Avec sous-domaines imbriqués
    - Nombre d'exigences par domaine
    """
    from sqlalchemy import text
    
    try:
        # Vérifier que le framework existe
        framework = db.query(Framework).filter_by(id=framework_id).first()
        if not framework:
            raise HTTPException(status_code=404, detail="Référentiel introuvable")
        
        # ✅ ÉTAPE 1 : Récupérer la hiérarchie des domaines (SANS COUNT)
        hierarchy_query = text("""
            WITH RECURSIVE domain_tree AS (
                -- Domaines racines (niveau 0)
                SELECT 
                    d.id,
                    d.code,
                    d.level,
                    d.parent_id,
                    d.sort_index,
                    COALESCE(dt.title, d.code) AS title,
                    ARRAY[d.id] AS path
                FROM domain d
                LEFT JOIN domain_title dt ON dt.domain_id = d.id 
                    AND dt.is_primary = true 
                    AND dt.language = 'fr'
                WHERE d.framework_id = :framework_id 
                    AND d.parent_id IS NULL
                
                UNION ALL
                
                -- Sous-domaines récursifs
                SELECT 
                    d.id,
                    d.code,
                    d.level,
                    d.parent_id,
                    d.sort_index,
                    COALESCE(dt.title, d.code) AS title,
                    dt_parent.path || d.id
                FROM domain d
                INNER JOIN domain_tree dt_parent ON d.parent_id = dt_parent.id
                LEFT JOIN domain_title dt ON dt.domain_id = d.id 
                    AND dt.is_primary = true 
                    AND dt.language = 'fr'
                WHERE d.framework_id = :framework_id
            )
            SELECT 
                id,
                code,
                title,
                level,
                parent_id,
                sort_index
            FROM domain_tree
            ORDER BY level, sort_index, code
        """)
        
        hierarchy_result = db.execute(hierarchy_query, {"framework_id": framework_id}).fetchall()
        
        # ✅ ÉTAPE 2 : Compter les exigences par domaine (requête séparée)
        count_query = text("""
            SELECT 
                domain_id,
                COUNT(*) AS requirement_count
            FROM requirement
            WHERE framework_id = :framework_id
            AND domain_id IS NOT NULL
            GROUP BY domain_id
        """)
        
        count_result = db.execute(count_query, {"framework_id": framework_id}).fetchall()
        
        # Créer un dictionnaire domain_id -> count
        requirement_counts = {str(row.domain_id): row.requirement_count for row in count_result}
        
        # ✅ ÉTAPE 3 : Construire la structure hiérarchique avec les counts
        domains_by_id = {}
        root_domains = []
        
        for row in hierarchy_result:
            domain_id = str(row.id)
            domain_dict = {
                "id": domain_id,
                "code": row.code,
                "title": row.title,
                "level": row.level,
                "requirement_count": requirement_counts.get(domain_id, 0),  # ✅ Fusion ici
                "children": []
            }
            domains_by_id[domain_id] = domain_dict
            
            if row.parent_id is None:
                root_domains.append(domain_dict)
            else:
                parent = domains_by_id.get(str(row.parent_id))
                if parent:
                    parent["children"].append(domain_dict)
        
        logger.info(f"✅ {len(root_domains)} domaine(s) racine(s) récupéré(s) pour {framework.code}")
        
        return root_domains
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erreur récupération domaines : {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@router.get(
    "/{framework_id}/requirements",
    response_model=List[Dict[str, Any]],
    summary="Récupérer les exigences d'un référentiel"
)
def get_framework_requirements(
    framework_id: str,
    domain: Optional[str] = None,
    section_id: Optional[str] = None,  # ✅ Alias pour domain (compatibilité frontend)
    limit: int = 1000,
    db: Session = Depends(get_db)
):
    """
    Récupère les exigences d'un référentiel, optionnellement filtrées par domaine.
    
    Args:
        framework_id: UUID du référentiel
        domain: UUID du domaine (optionnel)
        section_id: Alias pour domain (compatibilité frontend)
        limit: Nombre max de résultats
    
    Returns:
        Liste des exigences avec leurs métadonnées
    """
    try:
        # Vérifier que le framework existe
        framework = db.query(Framework).filter_by(id=framework_id).first()
        if not framework:
            raise HTTPException(status_code=404, detail="Référentiel introuvable")
        
        # ✅ CONSTRUCTION DYNAMIQUE DE LA REQUÊTE (avant text())
        
        # Utiliser section_id si domain n'est pas fourni (compatibilité)
        domain_id = domain or section_id
        
        # Base de la requête
        base_query = """
            WITH RECURSIVE domain_hierarchy AS (
                -- Niveau 0 : domaines racines
                SELECT 
                    d.id,
                    d.parent_id,
                    d.level,
                    COALESCE(
                        (SELECT dt.title 
                         FROM domain_title dt 
                         WHERE dt.domain_id = d.id 
                         AND dt.is_primary = true 
                         AND dt.language = 'fr' 
                         LIMIT 1),
                        d.code
                    ) AS domain_name,
                    ARRAY[d.id] AS path
                FROM domain d
                WHERE d.framework_id = :framework_id
                AND d.parent_id IS NULL
                
                UNION ALL
                
                -- Sous-domaines récursifs
                SELECT 
                    d.id,
                    d.parent_id,
                    d.level,
                    COALESCE(
                        (SELECT dt.title 
                         FROM domain_title dt 
                         WHERE dt.domain_id = d.id 
                         AND dt.is_primary = true 
                         AND dt.language = 'fr' 
                         LIMIT 1),
                        d.code
                    ) AS domain_name,
                    dh.path || d.id
                FROM domain d
                INNER JOIN domain_hierarchy dh ON d.parent_id = dh.id
            )
            SELECT 
                r.id,
                r.official_code,
                r.title,
                r.requirement_text,
                r.risk_level,
                r.created_at,
                dh.domain_name,
                dh.level,
                
                -- Récupérer le nom du domaine parent (niveau 0)
                (SELECT dh2.domain_name 
                 FROM domain_hierarchy dh2 
                 WHERE dh2.id = ANY(dh.path) 
                 AND dh2.level = 0 
                 LIMIT 1) AS root_domain
                
            FROM requirement r
            LEFT JOIN domain_hierarchy dh ON dh.id = r.domain_id
            WHERE r.framework_id = :framework_id
        """
        
        # ✅ Ajouter le filtre conditionnel
        params = {
            "framework_id": framework_id,
            "limit": limit
        }
        
        if domain_id:
            # ✅ Filtrer par domaine ET tous ses sous-domaines
            base_query += """
                AND r.domain_id IN (
                    SELECT id FROM domain_hierarchy 
                    WHERE :domain_id = ANY(path) OR id = :domain_id
                )
            """
            params["domain_id"] = domain_id
        
        base_query += """
            ORDER BY r.official_code NULLS LAST
            LIMIT :limit
        """
        
        # ✅ MAINTENANT on crée le text() avec la requête complète
        result = db.execute(text(base_query), params).fetchall()
        
        # ✅ Formater les résultats
        requirements = []
        for row in result:
            # Logique domain/subdomain :
            # - Si level = 0 ou None → domain uniquement
            # - Si level > 0 → domain = root, subdomain = actuel
            
            if row.level == 0 or row.level is None:
                domain_name = row.domain_name or "Non classé"
                subdomain_name = ""
            else:
                domain_name = row.root_domain or "Non classé"
                subdomain_name = row.domain_name or ""
            
            requirements.append({
                "id": str(row.id),
                "official_code": row.official_code or "",
                "title": row.title or "",
                "requirement_text": row.requirement_text or "",
                "domain": domain_name,
                "subdomain": subdomain_name,
                "risk_level": row.risk_level or "MEDIUM",
                "created_at": row.created_at.isoformat() if row.created_at else None
            })
        
        logger.info(f"✅ {len(requirements)} exigence(s) récupérée(s) pour {framework.code}")
        
        return requirements
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erreur récupération exigences : {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


